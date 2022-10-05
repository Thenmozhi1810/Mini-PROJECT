import cv2
from pyzbar.pyzbar import decode
from datetime import datetime
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os.path

cap = cv2.VideoCapture(0)
cap.set(3,640)
cap.set(4,480)

headers= ["Student Name","Register Number","Department","Stay Information","In Time"]
fname=str(datetime.now().date())+'.xlsx'
f_exists= os.path.exists(str(datetime.now().date())+'.xlsx')
scan=True
def markattendance(data):
    if f_exists==False:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    else:
        wb = load_workbook(fname)
        ws = wb.active
    time= datetime.now()
    intime = time.strftime('%H:%M:%S')
    student=[data[0],data[1],data[2],data[3],intime]
    ws.append(student)
    wb.save(filename=fname)
      
while scan ==True:
    sucess , img = cap.read()
    for barcode in decode(img):
        data = barcode.data.decode('utf-8')
        data="".join(data)
        data = data.replace('[','')
        data = data.replace(']','')
        data = data.replace("'","")
        data=data.split(',')
        scan=False
        markattendance(data)
        

    cv2.imshow("QRCode",img)
    cv2.waitKey(1)
